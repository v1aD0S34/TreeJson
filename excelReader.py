import openpyxl
from Models.Signal import Signal


# Читаем из Экселя и создаем список обьектов Signal
def read_and_save_signals(excel_settings_array, excel_file_path):
    signals_array = []
    workbook = openpyxl.load_workbook(excel_file_path)
    for setting in excel_settings_array:
        sheet = workbook[setting.ExcelSheet]
        row_count = sheet.max_row
        for i in range(setting.StartCount, row_count + setting.StartCount):
            if sheet.cell(row=i, column=1).value is not None:
                user_tree = (setting.TreePath + str(
                    sheet.cell(row=i, column=sheet[setting.Description + "10"].col_idx).value)).replace('/',
                                                                                                        '|').replace(
                    '\\', '/')
                opc_tag = setting.Prefix + "." + setting.Name + "." + str(
                    sheet.cell(row=i, column=sheet[setting.Tag + "10"].col_idx).value) + "." + setting.Postfix
                e_unit = str(sheet.cell(row=i, column=sheet[setting.Unit + "10"].col_idx).value).replace(
                    '\\', '/')
                description = str(sheet.cell(row=i, column=sheet[setting.Description + "10"].col_idx).value)
                _signal = Signal(user_tree, opc_tag, e_unit, description)
                signals_array.append(_signal)

    return signals_array
