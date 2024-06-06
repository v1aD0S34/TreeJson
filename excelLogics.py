import openpyxl
from Models.Signal import Signal
import os
from _publicConst import Const
from jsonLogics import get_data_from_config


# Читаем из Экселя и создаем список обьектов Signal
def read_and_save_signals(excel_settings_array, excel_file_path):
    signals_array = []
    workbook = openpyxl.load_workbook(excel_file_path)
    for setting in excel_settings_array:
        sheet = workbook[setting.ExcelSheet]
        row_count = sheet.max_row
        for i in range(setting.StartCount, row_count + setting.StartCount):
            if sheet.cell(row=i, column=1).value is not None:
                user_tree = (setting.TreePath + str(
                    sheet.cell(row=i, column=sheet[setting.Description + "10"].col_idx).value)).replace('/',
                                                                                                        '|').replace(
                    '\\', '/')
                opc_tag = setting.Prefix + "." + setting.Name + "." + str(
                    sheet.cell(row=i, column=sheet[setting.Tag + "10"].col_idx).value) + "." + setting.Postfix
                e_unit = str(sheet.cell(row=i, column=sheet[setting.Unit + "10"].col_idx).value).replace(
                    '\\', '/')
                description = str(sheet.cell(row=i, column=sheet[setting.Description + "10"].col_idx).value)
                _signal = Signal(user_tree, opc_tag, e_unit, description)
                signals_array.append(_signal)

    return signals_array


# Когда будет понимание про Идентификатор узла переписать
# Заполнение OПС-карты (екселя)
def process_excel_file(excel_name_opc, key_name):
    config = get_data_from_config(key_name)
    prefix_Alpha = get_data_from_config("prefix_Alpha") + "."
    prefix_Regul = get_data_from_config("prefix_Regul") + "."

    govno_BINDING = get_data_from_config("binding_OpcMap")
    govno_ADDRESS = get_data_from_config("addressSpace_OpcMap")
    govno_TYPE = get_data_from_config("typeId_OpcMap")
    input_file_path = Const.FILE_PATH + excel_name_opc
    output_file_path = Const.OUTPUT_FOLDER + excel_name_opc
    wb = openpyxl.load_workbook(filename=input_file_path)
    sheet = wb.active
    last_row = sheet.max_row

    for row_number in range(2, last_row + 1):
        a_value = sheet.cell(row=row_number, column=1).value
        if a_value:
            new_value, massIndex = process_tag_value(a_value, config, prefix_Alpha, prefix_Regul)
            sheet.cell(row=row_number, column=7, value=new_value)
            sheet.cell(row=row_number, column=8, value=massIndex)
            if new_value != "":
                sheet.cell(row=row_number, column=4, value=govno_BINDING)
                sheet.cell(row=row_number, column=5, value=govno_ADDRESS)
                sheet.cell(row=row_number, column=6, value=govno_TYPE)
            else:
                sheet.cell(row=row_number, column=4, value="не привязан")
                sheet.cell(row=row_number, column=5, value="")
                sheet.cell(row=row_number, column=6, value="")

    if not os.path.exists(Const.OUTPUT_FOLDER):
        os.makedirs(Const.OUTPUT_FOLDER)
    wb.save(output_file_path)

    output_file_absolute_path = os.path.abspath(output_file_path)
    if os.path.exists(output_file_path):
        print(f"Файл успешно создан по пути: {output_file_absolute_path}")
    else:
        print(f"Не удалось создать файл {output_file_absolute_path}")


def process_tag_value(tag_value, config, prefix_Alpha, prefix_Regul):
    for template in config:
        if (tag_value.startswith(prefix_Alpha + template["BeforeTag_StructAlpha"])
                and
                any(tag_value.endswith(signal) for signal in template["AfterTag_SignalAlpha"])
                and
                template["AfterTag_StructAlpha"] in tag_value):

            # Убираем начало и префикс
            tag = tag_value[len(prefix_Alpha + template["BeforeTag_StructAlpha"]):]

            # Находим конец
            before_signal_alpha = next(signal for signal in template["AfterTag_SignalAlpha"] if tag.endswith(signal))

            # tag = tag[: -len(before_signal_alpha)]  # Убираем коне
            tag = ((tag if len(before_signal_alpha) == 0 else tag[: -len(before_signal_alpha)])
                   .replace(template["AfterTag_StructAlpha"], ""))
            # tag = tag.translate(str.maketrans("", "", template["AfterTag_StructAlpha"]))
            # Получаем индекс совпавшего сигнала
            signal_index = template["AfterTag_SignalAlpha"].index(before_signal_alpha)

            # Проверяем размерности массивов и используем индексы соответственно
            before_signal_regul_index = signal_index if len(template["BeforeTag_SignalRegul"]) == len(
                template["AfterTag_SignalAlpha"]) else 0
            after_signal_regul_index = signal_index if len(template["AfterTag_SignalRegul"]) == len(
                template["AfterTag_SignalAlpha"]) else 0

            # Собираем новое значение с новыми префиксами и частями шаблона
            new_value = (prefix_Regul + template["BeforeTag_StructRegul"] +
                         template["BeforeTag_SignalRegul"][before_signal_regul_index] + tag +
                         template["AfterTag_SignalRegul"][after_signal_regul_index])
            massIndex = ""
            if template.get("IsArray", False):
                new_value, massIndex = check_string_with_number(new_value)
                # sheet.cell(row=row_number, column=8, value=massIndex)
            return new_value[: -1] if new_value.endswith('.') else new_value, massIndex

    return "", ""


def check_string_with_number(input_string):
    # Проверяем, заканчивается ли строка на число
    if input_string[-1].isdigit():
        # Находим индекс последней цифры
        for i in range(len(input_string) - 1, -1, -1):
            if not input_string[i].isdigit():
                break
        # Извлекаем число
        number = int(input_string[i + 1:])
        # Проверяем, есть ли точка перед числом
        # if input_string[i] == '.':
        #     return input_string[:i], number
        # else:
        #     return input_string, number
        return input_string[: i + 1], number
    else:
        # return input_string, ""
        return input_string, 0 if any(
            key in input_string for key in ["StrStep", "StrColor", "StrText"]) else ""